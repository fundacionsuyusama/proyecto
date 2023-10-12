from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import AuthenticationForm
from .models import *
from .forms import *
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.contrib.auth.models import User

@login_required(login_url='user_login')
def progreso(request):
    resultados = Resultado.objects.all()
    total_promedio = 0
    cantidad_resultados = 0
    
    for resultado in resultados:
        actividades = resultado.actividad_set.all()
        total_avance = 0
        cantidad_actividades = 0
        
        for actividad in actividades:
            avances = actividad.avance_set.all()
            total_actividad = sum(avance.contenido for avance in avances)
            cantidad_actividades += 1
            total_avance += total_actividad
        
        if cantidad_actividades > 0:
            promedio_avance = round(total_avance / cantidad_actividades)
        else:
            promedio_avance = 0
        
        resultado.promedio_avance = promedio_avance
        total_promedio += promedio_avance
        cantidad_resultados += 1
    
    if cantidad_resultados > 0:
        promedio_total = round(total_promedio / cantidad_resultados)
    else:
        promedio_total = 0
    
    return render(request, 'main/tiempo/progreso.html', {'resultados': resultados, 'promedio_total': promedio_total})

@login_required(login_url='user_login')
def exportar_datos(request):
    variable = "Hola desde exportar"
    return render(request, 'main/herramientas/exportar.html')


@login_required(login_url='user_login')
def exportar(request):
    # Obtén el usuario actual
    usuario_actual = request.user

    # Crea un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"

    # Definir la lista de campos que deseas exportar
    campos_exportar = [
        'Nombre del Proyecto', 'Nombre del Resultado', 'Texto del Resultado',
        'Nombre de la Actividad', 'Contenido de la Actividad', 'Fecha de Vencimiento de la Actividad',
        'Contenido del Avance', 'Cumplido', 'Proceso', 'Urgente',
        'Contenido de la Dificultad', 'Contenido de la Alternativa',
        'Nombre de la Seccion', 'Contenido de la Seccion', 'Avance de la Seccion'
    ]

    for col_num, encabezado in enumerate(campos_exportar, 1):
        col_letra = get_column_letter(col_num)
        ws[f'{col_letra}1'] = encabezado

    # Aplica formato a los encabezados
    for cell in ws['1:1']:
        cell.font = Font(bold=True)

    # Obtiene los proyectos del usuario actual
    proyectos = Proyecto.objects.filter(user=usuario_actual)

    row_num = 2
    for proyecto in proyectos:
        # Agrega el nombre del proyecto
        ws.cell(row=row_num, column=1, value=proyecto.nombre)

        resultados = Resultado.objects.filter(proyecto=proyecto)

        if resultados.exists():
            for resultado in resultados:
                ws.cell(row=row_num, column=2, value=resultado.nombre)
                ws.cell(row=row_num, column=3, value=resultado.texto)
                row_num += 1

                actividades = Actividad.objects.filter(resultado=resultado)

                for actividad in actividades:
                    ws.cell(row=row_num, column=4, value=actividad.nombre)
                    ws.cell(row=row_num, column=5, value=actividad.contenido)

                    fecha_vencimiento = actividad.fecha_vencimiento
                    if fecha_vencimiento:
                        fecha_vencimiento = fecha_vencimiento.replace(tzinfo=None)
                    ws.cell(row=row_num, column=6, value=fecha_vencimiento)

                    # Obtener los avances de la actividad
                    avances = Avance.objects.filter(actividad=actividad)

                    if avances.exists():
                        # Recorrer los avances y agregarlos
                        avances_contenido = ', '.join(str(avance.contenido) for avance in avances)
                        ws.cell(row=row_num, column=7, value=avances_contenido)
                    else:
                        ws.cell(row=row_num, column=7, value="")

                    # Obtener los Cumplidos de la actividad
                    cumplidos = Cumplido.objects.filter(actividad=actividad)
                    procesos = Proceso.objects.filter(actividad=actividad)
                    urgentes = Urgente.objects.filter(actividad=actividad)

                    # Agregar los Cumplidos en celdas separadas
                    if cumplidos.exists():
                        cumplidos_contenido = ', '.join(cumplido.cumplida for cumplido in cumplidos)
                        ws.cell(row=row_num, column=8, value=cumplidos_contenido)
                    else:
                        ws.cell(row=row_num, column=8, value="")

                    # Agregar los Procesos en celdas separadas
                    if procesos.exists():
                        procesos_contenido = ', '.join(proceso.proceso for proceso in procesos)
                        ws.cell(row=row_num, column=9, value=procesos_contenido)
                    else:
                        ws.cell(row=row_num, column=9, value="")

                    # Agregar los Urgentes en celdas separadas
                    if urgentes.exists():
                        urgentes_contenido = ', '.join(urgente.urgente for urgente in urgentes)
                        ws.cell(row=row_num, column=10, value=urgentes_contenido)
                    else:
                        ws.cell(row=row_num, column=10, value="")

                    # Obtener las dificultades de la actividad
                    dificultades = Dificultad.objects.filter(actividad=actividad)
                    alternativas = []

                    for dificultad in dificultades:
                        alternativa = Alternativa.objects.filter(dificultad=dificultad).first()
                        alternativas.append(alternativa)

                    if dificultades.exists():
                        # Recorrer las dificultades y agregarlas en celdas separadas
                        for i, dificultad in enumerate(dificultades):
                            dificultad_contenido = dificultad.contenido
                            alternativa_contenido = alternativas[i].contenido if alternativas[i] else ""
                            ws.cell(row=row_num, column=11, value=dificultad_contenido)
                            ws.cell(row=row_num, column=12, value=alternativa_contenido)
                            row_num += 1
                    else:
                        ws.cell(row=row_num, column=11, value="")
                        ws.cell(row=row_num, column=12, value="")
                        row_num += 1

                    # Obtener las secciones de la actividad
                    secciones = Seccion.objects.filter(actividad=actividad)

                    for seccion in secciones:
                        ws.cell(row=row_num, column=13, value=seccion.nombre)
                        ws.cell(row=row_num, column=14, value=seccion.contenido)
                        ws.cell(row=row_num, column=15, value=seccion.avance)
                        row_num += 1
        else:
            for col_num in range(2, len(campos_exportar) + 1):
                ws.cell(row=row_num, column=col_num, value="")
            row_num += 1

    # Configura la respuesta HTTP para devolver el archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{usuario_actual.username}_proyectos.xlsx"'

    # Guarda el libro de Excel en la respuesta
    wb.save(response)

    return response


#! En Urgente - Análisis tiempo
@login_required(login_url='user_login')
def eliminar_urgente(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    urgente = get_object_or_404(Urgente, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        urgente.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)
    
    return render(request, 'main/tiempo/urgente/eliminar-urgente.html', {'urgente': urgente, 'resultado': resultado, ' resultado_id': resultado_id, 'proyecto': proyecto})

@login_required(login_url='user_login')
def editar_urgente(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    urgente = get_object_or_404(Urgente, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        urgente.urgente = request.POST.get('urgente')
        urgente.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id,)
    
    return render(request, 'main/tiempo/urgente/editar-urgente.html', {'urgente': urgente, 'resultado': resultado, 'proyecto': proyecto})

@login_required(login_url='user_login')
def crear_urgente(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        urgente = request.POST['urgente']
        tiempo = actividad.urgente_set.create(urgente=urgente)
        tiempo.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/tiempo/urgente/urgente.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto})

#! En proceso - Análisis tiempo
@login_required(login_url='user_login')
def eliminar_proceso(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proceso = get_object_or_404(Proceso, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        proceso.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)
    
    return render(request, 'main/tiempo/proceso/eliminar-proceso.html', {'proceso': proceso, 'resultado': resultado, ' resultado_id': resultado_id, 'proyecto': proyecto})

@login_required(login_url='user_login')
def editar_proceso(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proceso = get_object_or_404(Proceso, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        proceso.proceso = request.POST.get('proceso')
        proceso.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id,)
    
    return render(request, 'main/tiempo/proceso/editar-proceso.html', {'proceso': proceso, 'resultado': resultado, 'proyecto': proyecto})

@login_required(login_url='user_login')
def crear_proceso(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        proceso = request.POST['proceso']
        tiempo = actividad.proceso_set.create(proceso=proceso)
        tiempo.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/tiempo/proceso/proceso.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto})

#! Cumplida - Análisis tiempo
@login_required(login_url='user_login')
def eliminar_cumplida(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    cumplido = get_object_or_404(Cumplido, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        cumplido.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)
    
    return render(request, 'main/tiempo/cumplida/eliminar-cumplida.html', {'cumplido': cumplido, 'resultado': resultado, ' resultado_id': resultado_id, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def editar_cumplida(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    cumplido = get_object_or_404(Cumplido, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        cumplido.cumplida = request.POST.get('cumplida')
        cumplido.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id,)
    
    return render(request, 'main/tiempo/cumplida/editar-cumplida.html', {'cumplido': cumplido, 'resultado': resultado, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def crear_cumplida(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        cumplida = request.POST['cumplida']
        cumplido = actividad.cumplido_set.create(cumplida=cumplida)
        cumplido.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/tiempo/cumplida/cumplida.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto,})

#! Alternativa
@login_required(login_url='user_login')
def eliminar_alternativa(request, resultado_id, actividad_id, dificultad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    alternativa = get_object_or_404(Alternativa, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        alternativa.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)
    
    return render(request, 'main/alternativa/eliminar_alternativa.html', {'alternativa': alternativa, 'resultado': resultado, 'resultado_id': resultado_id, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def editar_alternativa(request, resultado_id, actividad_id, dificultad_id, id, proyecto_id):
    alternativa = get_object_or_404(Alternativa, id=id)
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        alternativa.contenido = request.POST.get('contenido')
        alternativa.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)

    return render(request, 'main/alternativa/editar_alternativa.html', {'alternativa': alternativa, 'resultado': resultado, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def crear_alternativa(request, resultado_id, actividad_id, dificultad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    dificultad = get_object_or_404(Dificultad, id=dificultad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        contenido = request.POST['contenido']
        alternativa = dificultad.alternativa_set.create(contenido=contenido)
        alternativa.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/alternativa/crear_alternativa.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto,})

#! Dificultad
@login_required(login_url='user_login')
def eliminar_dificultad(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    dificultad = get_object_or_404(Dificultad, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        dificultad.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)
    
    return render(request, 'main/dificultad/eliminar_dificultad.html', {'dificultad': dificultad, 'resultado':resultado, 'resultado_id': resultado_id, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def editar_dificultad(request, resultado_id, actividad_id, id, proyecto_id):
    dificultad = get_object_or_404(Dificultad, id=id)
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        dificultad.contenido = request.POST.get('contenido')
        dificultad.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id,)
    
    return render(request, 'main/dificultad/editar_dificultad.html', {'dificultad': dificultad, 'resultado': resultado, 'proyecto': proyecto})

@login_required(login_url='user_login')
def crear_dificultad(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        contenido = request.POST['contenido']
        dificultad = actividad.dificultad_set.create(contenido=contenido)
        dificultad.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/dificultad/crear_dificultad.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto,})

#! Avance
@login_required(login_url='user_login')
def eliminar_avance(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    avance = get_object_or_404(Avance, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        avance.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado_id)
    
    return render(request, 'main/avance/eliminar_avance.html', {'avance': avance, 'resultado': resultado, ' resultado_id': resultado_id, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def editar_avance(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    avance = get_object_or_404(Avance, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    
    if request.method == 'POST':
        avance.contenido = request.POST.get('contenido')
        avance.save()
        return redirect('ver_actividad',  proyecto_id=proyecto.id, resultado_id=resultado_id,)
    
    return render(request, 'main/avance/editar_avance.html', {'avance': avance, 'resultado': resultado, 'proyecto': proyecto})

@login_required(login_url='user_login')
def crear_avance(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        contenido = request.POST['contenido']
        avance = actividad.avance_set.create(contenido=contenido)
        avance.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/avance/crear_avance.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto,})

#! Seccion
@login_required(login_url='user_login')
def guardar_seccion(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    if request.method == 'POST':
        for key, value in request.POST.items():
            if key.startswith('seccion_'):
                seccion_id = key.split('_')[-1]
                seccion_status = value
                try:
                    seccion = Seccion.objects.get(id=seccion_id)
                    seccion.is_completed = (seccion_status == "completed")
                    seccion.save()
                except Seccion.DoesNotExist:
                    pass

    return redirect('ver_secciones', proyecto_id=proyecto_id, resultado_id=resultado_id, actividad_id=actividad_id)

@login_required(login_url='user_login')
def ver_secciones(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    secciones = actividad.seccion_set.all()
    total_secciones = secciones.count()
    secciones_terminadas = Seccion.objects.filter(actividad=actividad)

    # Código para las secciones de la fecha de vencimiento
    for seccion_datos in secciones_terminadas:
        seccion_datos.fecha_vencimiento_minus_24 = seccion_datos.fecha_vencimiento_seccion - timedelta(hours=24)

    actividad_actual = actividad.fecha_vencimiento

    current_datetime = timezone.now()
    tiempo_real =  actividad_actual - current_datetime

    horas_restantes = tiempo_real.seconds // 3600
    minutos_restantes = (tiempo_real.seconds % 3600) // 60

    i = 0

    if tiempo_real.days == 0:
        if horas_restantes == 0 and minutos_restantes == 0:
            mensaje = "No queda tiempo"
        elif horas_restantes == 0:
            if minutos_restantes == 1:
                mensaje = "1 minuto"
            else:
                mensaje = f"{minutos_restantes} minutos"
        elif minutos_restantes == 0:
            mensaje = f"{horas_restantes} horas"
        elif horas_restantes == 1:
            mensaje = f"1 hora y {minutos_restantes} minutos"
        else:
            if minutos_restantes == 1:
                mensaje = f"{horas_restantes} horas y 1 minuto"
            else:
                mensaje = f"{horas_restantes} horas y {minutos_restantes} minutos"
    else:
        mensaje = 0;


    secciones_completadas = sum(seccion.is_completed for seccion in secciones)
    

    if total_secciones > 0:
        porcentaje = round((secciones_completadas / total_secciones) * 100)
        total_porcentaje = round(100 / total_secciones)

        dias_restantes = tiempo_real.days
    
        divisor = dias_restantes // total_secciones
        residuo = divisor + (dias_restantes % total_secciones)

        variables = [divisor] * total_secciones

        for i in range(dias_restantes % total_secciones):
            variables[i] += 1

    else:
        porcentaje = 0
        total_porcentaje = 0
        divisor = 0
        variables = 0

    
    return render(request, 'main/secciones/ver_secciones.html', {'resultado': resultado, 'actividad': actividad, 'secciones': secciones, 'actividad_id': actividad_id, 'total_secciones': total_secciones, 'secciones_completadas': secciones_completadas, 'porcentaje': porcentaje, 'resultado_id': resultado_id, 'total_porcentaje': total_porcentaje, 'actividad_actual': actividad_actual, 'tiempo_real': tiempo_real, 'mensaje': mensaje, 'variables': variables, 'current_datetime': current_datetime, 'secciones_terminadas': secciones_terminadas, 'horas_restantes': horas_restantes, 'proyecto': proyecto})

@login_required(login_url='user_login')
def eliminar_secciones(request, resultado_id, actividad_id, id, proyecto_id):
    seccion = get_object_or_404(Seccion, id=id)
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        seccion.delete()
        return redirect('ver_secciones', proyecto_id=proyecto.id, resultado_id=resultado_id, actividad_id=actividad_id)

    return render(request, 'main/secciones/eliminar_secciones.html', {'seccion': seccion, 'resultado': resultado, 'actividad_id': actividad_id, 'proyecto': proyecto})

@login_required(login_url='user_login')
def editar_secciones(request, resultado_id, actividad_id, id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    seccion = get_object_or_404(Seccion, id=id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        seccion.nombre = request.POST.get('nombre')
        seccion.contenido = request.POST.get('contenido')
        seccion.avance = request.POST.get('avance')
        seccion.fecha_vencimiento_seccion = request.POST.get('fecha_vencimiento_seccion')
        seccion.save()
        return redirect('ver_secciones', proyecto_id=proyecto.id, resultado_id=resultado_id, actividad_id=actividad_id)

    return render(request, 'main/secciones/editar_secciones.html', {'seccion': seccion, 'resultado': resultado, 'actividad_id': actividad_id, 'proyecto': proyecto})

@login_required(login_url='user_login')
def crear_secciones(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        nombre = request.POST['nombre']
        avance = request.POST['avance']
        contenido = request.POST['contenido']
        fecha_vencimiento_seccion = request.POST['fecha_vencimiento_seccion']

        seccion = actividad.seccion_set.create(
            nombre=nombre,
            avance=avance,
            contenido=contenido,
            fecha_vencimiento_seccion=fecha_vencimiento_seccion
            )
        seccion.save()
        return redirect('ver_secciones', proyecto_id=proyecto.id, resultado_id=resultado_id, actividad_id=actividad_id)

    return render(request, 'main/secciones/crear_secciones.html', {'actividad': actividad, 'resultado': resultado, 'actividad_id': actividad_id, 'proyecto': proyecto})

#! Actividad
@login_required(login_url='user_login')
def eliminra_actividad(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        actividad.delete()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    return render(request, 'main/actividad/eliminar_actividad.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto,})

@login_required(login_url='puruser_login')
def editar_actividad(request, resultado_id, actividad_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    actividad = get_object_or_404(Actividad, id=actividad_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        actividad.nombre = request.POST.get('nombre')
        actividad.contenido = request.POST.get('contenido')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        actividad.fecha_vencimiento = timezone.make_aware(datetime.strptime(fecha_vencimiento, '%Y-%m-%dT%H:%M'))
        actividad.save()
        
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/actividad/editar_actividad.html', {'actividad': actividad, 'resultado': resultado, 'proyecto': proyecto})

@login_required(login_url='user_login')
def crear_actividad(request, resultado_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    

    if request.method == 'POST':
        nombre = request.POST['nombre']
        contenido = request.POST['contenido']
        fecha_vencimiento = request.POST['fecha_vencimiento']

        actividad = resultado.actividad_set.create(
            nombre=nombre,
            contenido=contenido,
            fecha_vencimiento=fecha_vencimiento,
            fecha_actual=timezone.now()
        )
        actividad.save()
        return redirect('ver_actividad', proyecto_id=proyecto.id, resultado_id=resultado.id)
    
    return render(request, 'main/actividad/crear_actividad.html', {'resultado': resultado, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def ver_actividad(request, resultado_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    actividades = resultado.actividad_set.all().order_by('id')

    

    current_datetime = timezone.now()

    return render(request, 'main/actividad/ver_actividad.html', {'resultado': resultado, 'actividades': actividades, 'current_datetime': current_datetime, 'proyecto': proyecto,})

#! Resultado
@login_required(login_url='user_login')
def eliminar_resultado(request, resultado_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    if request.method == 'POST':
        resultado.delete()
        return redirect('ver_resultados', proyecto_id=proyecto.id)
    return render(request, 'main/resultado/eliminar_resultado.html', {'resultado': resultado, 'proyecto': proyecto})

@login_required(login_url='user_login')
def modificar_resultado(request, resultado_id, proyecto_id):
    resultado = get_object_or_404(Resultado, id=resultado_id)
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        resultado.nombre = request.POST.get('nombre')
        resultado.texto = request.POST.get('texto')
        resultado.save()
        return redirect('ver_resultados', proyecto_id=proyecto.id)
    return render(request, 'main/resultado/modificar_resultado.html', {'resultado': resultado, 'proyecto': proyecto,})

@login_required(login_url='user_login')
def crear_resultado(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    fecha_actual = timezone.now()

    if request.method == 'POST':
        nombre = request.POST['nombre']
        texto = request.POST['texto']

        resultado = Resultado.objects.create(proyecto=proyecto, nombre=nombre, texto=texto)
        return redirect('ver_resultados', proyecto_id=proyecto.id)
    
    return render(request, 'main/resultado/crear_resultado.html', {'fecha_actual': fecha_actual, 'proyecto': proyecto})

@login_required(login_url='user_login')
def ver_resultados(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    resultados = proyecto.resultados.all()
    total_promedio = 0
    cantidad_resultados = 0
    
    for resultado in resultados:
        actividades = resultado.actividad_set.all()
        total_avance = 0
        cantidad_actividades = 0
        
        for actividad in actividades:
            avances = actividad.avance_set.all()
            total_actividad = sum(avance.contenido for avance in avances)
            cantidad_actividades += 1
            total_avance += total_actividad
        
        if cantidad_actividades > 0:
            promedio_avance = round(total_avance / cantidad_actividades)
        else:
            promedio_avance = 0
        
        resultado.promedio_avance = promedio_avance
        total_promedio += promedio_avance
        cantidad_resultados += 1
    
    if cantidad_resultados > 0:
        promedio_total = round(total_promedio / cantidad_resultados)
    else:
        promedio_total = 0
    return render(request, 'main/resultado/resultado2.html', {'proyecto': proyecto, 'resultados': resultados, 'promedio_total':promedio_total})

#! Proyecto
@login_required(login_url='user_login')
def eliminar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        proyecto.delete()
        return redirect('home')

    return render(request, 'main/proyecto/eliminar_proyecto.html', {'proyecto': proyecto,})

@login_required(login_url='user_login')
def editar_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    if request.method == 'POST':
        proyecto.nombre = request.POST.get('nombre')
        proyecto.save()
        return redirect('home')

    return render(request, 'main/proyecto/editar_proyecto.html', {'proyecto': proyecto,})

@login_required(login_url='user_login')
def crear_proyecto(request):
    
    if request.method == 'POST':
        nombre = request.POST['nombre']
        proyecto = Proyecto.objects.create(user=request.user, nombre=nombre)
        return redirect('home')

    return render(request, 'main/proyecto/crear_proyecto.html')

@login_required(login_url='user_login')
def home(request):
    # Filtra los proyectos del usuario actual
    proyectos_del_usuario = Proyecto.objects.filter(user=request.user)

    return render(request, 'main/proyecto/home.html', {'proyectos': proyectos_del_usuario})

@login_required(login_url='user_login')
def index(request):
    return render(request, 'index.html')

def user_login(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'GET':
        return render(request, 'login/login.html', {'form': AuthenticationForm})
    else:
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'login/login.html', {'form': AuthenticationForm,'error': 'El nombre de usuario o contraseña son incorrectos'})
        else:
            login(request, user)
            return redirect('home')
    

